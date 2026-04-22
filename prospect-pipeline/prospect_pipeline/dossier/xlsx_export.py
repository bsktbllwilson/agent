"""Weekly XLSX with five tabs:

    - Hot Prospects      — high confidence + work contact resolved
    - Warm Leads         — named but missing contact, or medium confidence
    - Indirect Outreach  — unresolved buyers with attorney / managing-agent path
    - Contract Signed    — UrbanDigs contract-signed events with enrichment
    - Holdings Patterns  — buyers with holdings_count >= 2
"""
from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .csv_export import CSV_COLUMNS


def _bucket(rows: Sequence[dict]) -> dict[str, list[dict]]:
    hot, warm, indirect, contracts, holdings = [], [], [], [], []
    for r in rows:
        has_contact = bool(r.get("work_email") or r.get("work_phone"))
        is_contract = r.get("last_trigger_event_id") and (r.get("resolution_path") or "").count("tier") == 0 and r.get("recommended_channel") == "selling_broker" and (r.get("source") == "urbandigs" or "urbandigs" in (r.get("resolution_path") or ""))

        if r.get("confidence") == "high" and has_contact:
            hot.append(r)
        elif r.get("confidence") in ("high", "medium"):
            warm.append(r)
        else:
            indirect.append(r)

        if r.get("holdings_count") and r["holdings_count"] >= 1:
            holdings.append(r)
    return {
        "Hot Prospects": hot,
        "Warm Leads": warm,
        "Indirect Outreach": indirect,
        "Contract Signed": contracts,   # filled via write_xlsx
        "Holdings Patterns": holdings,
    }


def _flatten(v):
    if isinstance(v, list):
        return "; ".join(
            str(item.get("address") if isinstance(item, dict) else item)
            for item in v
        )
    return v if v is not None else ""


def _write_sheet(wb: Workbook, title: str, rows: Sequence[dict]) -> None:
    ws = wb.create_sheet(title=title[:31])
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for i, col in enumerate(CSV_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
    for r_idx, r in enumerate(rows, start=2):
        for c_idx, col in enumerate(CSV_COLUMNS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_flatten(r.get(col)))
    # Reasonable column widths
    widths = {
        "prospect_id": 22, "legal_name": 28, "resolved_principal": 26,
        "confidence": 10, "resolution_path": 30, "buyer_type": 18,
        "primary_firm": 28, "title": 24, "estimated_net_worth_band": 14,
        "known_affiliations": 36, "geography_signal": 14,
        "other_nyc_holdings": 40, "holdings_count": 8,
        "work_email": 32, "work_phone": 18, "linkedin_url": 40,
        "assistant_contact": 26, "attorney_firm": 28, "managing_agent": 22,
        "broker_to_contact": 22, "recommended_channel": 18,
        "outreach_angle": 50, "sensitivity_flags": 26,
        "last_trigger_event_id": 18, "first_seen_date": 20,
        "last_contacted_date": 20, "outreach_status": 14,
    }
    for i, col in enumerate(CSV_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 18)
    ws.freeze_panes = "A2"


def write_xlsx(prospects: Sequence[dict], contract_rows: Sequence[dict], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Replace default sheet
    default = wb.active
    wb.remove(default)
    buckets = _bucket(prospects)
    buckets["Contract Signed"] = list(contract_rows)
    for title in ("Hot Prospects", "Warm Leads", "Indirect Outreach", "Contract Signed", "Holdings Patterns"):
        _write_sheet(wb, title, buckets.get(title, []))
    wb.save(path)
    return path
