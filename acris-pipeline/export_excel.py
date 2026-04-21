"""Weekly Excel snapshot of recent ACRIS comps from Airtable."""

import argparse
import os
import sys
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pyairtable import Api

COLUMNS = [
    "recorded_date",
    "doc_date",
    "borough",
    "address",
    "bbl",
    "property_type",
    "price",
    "sqft",
    "ppsf",
    "all_cash_flag",
    "llc_flag",
    "trust_flag",
    "entity_type",
    "beneficial_owner_category",
    "buyer_name",
    "seller_name",
    "acris_url",
]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
COLUMN_WIDTH = 18


def fetch_records(days: int) -> list[dict]:
    api_key = os.environ["AIRTABLE_API_KEY"]
    base_id = os.environ["AIRTABLE_BASE_ID"]
    table_name = os.environ["AIRTABLE_TABLE_NAME"]
    cutoff = (date.today() - timedelta(days=days)).isoformat()
    formula = f"IS_AFTER({{recorded_date}}, '{cutoff}')"
    api = Api(api_key)
    table = api.table(base_id, table_name)
    return table.all(formula=formula)


def write_xlsx(records: list[dict], out_path: str) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "comps"

    ws.append(COLUMNS)
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[cell.column_letter].width = COLUMN_WIDTH

    for rec in records:
        fields = rec.get("fields", {})
        ws.append([fields.get(c) for c in COLUMNS])

    wb.save(out_path)
    return len(records)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Export recent comps to Excel")
    parser.add_argument("--days", type=int, default=30)
    parser.add_argument(
        "--out",
        default=f"comps_{date.today().strftime('%Y%m%d')}.xlsx",
    )
    args = parser.parse_args(argv)

    records = fetch_records(args.days)
    n = write_xlsx(records, args.out)
    print(f"Wrote {n} rows → {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
